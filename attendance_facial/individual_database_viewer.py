import sys

sys.path.append('~/.local/lib/python3.5/site-packages')                                     #for ros users
sys.path.remove('/opt/ros/kinetic/lib/python2.7/dist-packages')                             #for ros users
file_name ='/home/sammy-ros/attendance_facial//attendance.xlsx'
import os
import tkinter as tk
import tkinter.font as font
from openpyxl import Workbook,load_workbook

file_name ='/home/sammy-ros/attendance_facial//attendance.xlsx'
if os.path.exists(file_name):
    wb = load_workbook(filename = file_name)
    sh1 = wb.active

Roll_No = []
Name  = []

for values in sh1.iter_cols(min_row = 2, min_col  = 1, max_col = 1, values_only = True):
    for x in values:
        if x != None : Roll_No.append(x)
for values in sh1.iter_cols(min_row = 2, min_col  = 2, max_col = 2, values_only = True):
    for x in values:
        if x != None : Name.append(x)
 


def load_database():
	if entry_1.get() in Name:
		index = Name.index(entry_1.get())+2
	elif entry_1.get() in Roll_No:
		index = Roll_No.index(entry_1.get())+2
	else:
		entry_1.delete(0,tk.END)
		return 0
	global label_2,label_3,label_4,label_5,label_6,label_7,label_8

	print(index)

	disp_name = sh1.cell(row = index, column = 2).value 
	disp_roll = sh1.cell(row = index, column = 1).value



	label_2.grid_forget()
	label_3.grid_forget()
	label_4.grid_forget()
	label_5.grid_forget()
	label_6.grid_forget()
	label_7.grid_forget()
	label_8.grid_forget()



	label_2 = tk.Label(master = frame_3, text = "Name:\n"+ disp_name, relief = tk.RAISED,  bg = "lightblue", fg = 'red')
	label_2['font'] = font.Font(size = 15)
	label_2.grid(row = 0, column = 0,sticky = "news", padx = 10)


	label_3 = tk.Label(master = frame_3, text = "Roll Number:\n"+disp_roll,  relief = tk.RAISED, bg = "lightblue", fg = 'red')
	label_3['font'] = font.Font(size = 15)
	label_3.grid(row = 0, column = 1 ,sticky = "news", padx = 10)

	for values in sh1.iter_rows(min_row = 1, max_row =1, values_only = True):
		disp_wrk  = len(values) -2

	label_4 = tk.Label(master = frame_3,relief = tk.RAISED, text = "Total No of working Days:\n"+str(disp_wrk), bg = "lightblue", fg = 'red')
	label_4['font'] = font.Font(size = 15)
	label_4.grid(row = 1, column = 0 ,sticky = "news", padx = 10)

	att_data= []

	for values in sh1.iter_cols(min_row = index, min_col  = 3, max_row = index, values_only = True):
	    for x in values:
	        if x != None : att_data.append(x)
	disp_abs = att_data.count("A")
	disp_pre = disp_wrk - disp_abs
	disp_perc = disp_pre*100/disp_wrk

	label_5 = tk.Label(master = frame_3, relief = tk.RAISED, text = "Attendance Percentage:\n"+str(disp_perc)+" %", bg = "lightblue", fg = 'red')
	label_5['font'] = font.Font(size = 15)
	label_5.grid(row = 1, column = 1,sticky = "news", padx = 10)

	label_6 = tk.Label(master = frame_3, relief = tk.RAISED, text = "No of days present:\n"+str(disp_pre), bg = "lightblue", fg = 'red')
	label_6['font'] = font.Font(size = 15)
	label_6.grid(row = 2, column = 0,sticky = "news", padx = 10)


	label_7 = tk.Label(master = frame_3,relief = tk.RAISED, text = "No of days Absent:\n"+str(disp_abs), bg = "lightblue", fg = 'red')
	label_7['font'] = font.Font(size = 15)
	label_7.grid(row = 2, column = 1 ,sticky = "news", padx = 10)


	if disp_perc>75:

		label_8 = tk.Label(master = frame_4,relief = tk.RIDGE, text = disp_name+" is Eligible for the upcoming Examinations",bg = "white", fg = 'green')
		label_8['font'] = font.Font(size = 15)
		label_8.grid(row = 0, column = 0 ,sticky = "news", padx = 50)



	else:

		label_8 = tk.Label(master = frame_4,relief = tk.RIDGE, text = disp_name+" is not Eligible for the upcoming Examinations",bg = "white", fg = 'red')
		label_8['font'] = font.Font(size = 15)
		label_8.grid(row = 0, column = 0 ,sticky = "news", padx = 50)



























window = tk.Tk()
window.title('Individual Database Viewer')


frame_1 = tk.Frame(window,width  = 900, height = 100 , bg = "lightgreen")
frame_1.pack()


frame_1.rowconfigure(0,weight = 1 , minsize = 25)
frame_1.columnconfigure(0,weight = 1 , minsize = 1000)


label = tk.Label(master = frame_1, text = "Individual Database Viewer", bg = "lightgreen", fg = 'red')
label['font'] = font.Font(family = 'Times',size = 30)
label.grid(row = 0, column = 0,sticky = "news", padx = 5)


frame_2 = tk.Frame(window,width  = 1900, height = 100 , bg = "lightblue")
frame_2.pack()

frame_2.rowconfigure(0,weight = 1 , minsize = 30)

for x in range(2):
	frame_2.columnconfigure(x ,weight = 1 , minsize = 300)
frame_2.columnconfigure(3 ,weight = 1 , minsize = 168)


		
label_1 = tk.Label(master = frame_2, text = "Enter the Name or Roll_No:", bg = "lightblue", fg = 'red')
label_1['font'] = font.Font(size = 15)
label_1.grid(row = 0, column = 0,sticky = "news", padx = 50)


entry_1 = tk.Entry(master = frame_2 )
entry_1['font'] = font.Font(size = 12)
entry_1.grid(row = 0, column = 1,sticky = "news",padx = 50, pady = 15)


load_button = tk.Button(master = frame_2, text = 'Load',bg = "lightblue" , fg = "blue" ,command =  load_database)
load_button['font'] = font.Font(size = 15)
load_button.grid(row=  0,column = 2, sticky = "nsew", padx = 35, pady = 15)

frame_3 = tk.Frame(window,width  = 900, height = 100 , bg = "lightblue")
frame_3.pack()


for x in range(2):
	frame_3.columnconfigure(x ,weight = 1 , minsize = 500)
for x in range(3):frame_3.rowconfigure(x,weight = 1 , minsize = 70)

global label_2#,label_3,label_4,label_5,label_6,label_7,label_8

label_2 = tk.Label(master = frame_3, relief = tk.RAISED, text = "Name:\n\t", bg = "lightblue", fg = 'red')
label_2['font'] = font.Font(size = 15)
label_2.grid(row = 0, column = 0,sticky = "news", padx = 10)

label_3 = tk.Label(master = frame_3,relief = tk.RAISED, text = "Roll Number:\n\t", bg = "lightblue", fg = 'red')
label_3['font'] = font.Font(size = 15)
label_3.grid(row = 0, column = 1 ,sticky = "news", padx = 10)


label_4 = tk.Label(master = frame_3,relief = tk.RAISED, text = "Total No of working Days:\n\t", bg = "lightblue", fg = 'red')
label_4['font'] = font.Font(size = 15)
label_4.grid(row = 1, column = 0 ,sticky = "news", padx = 10)

label_5 = tk.Label(master = frame_3, relief = tk.RAISED, text = "Attendance Percentage:\n\t", bg = "lightblue", fg = 'red')
label_5['font'] = font.Font(size = 15)
label_5.grid(row = 1, column = 1,sticky = "news", padx = 10)

label_6 = tk.Label(master = frame_3, relief = tk.RAISED, text = "No of days present:\n\t", bg = "lightblue", fg = 'red')
label_6['font'] = font.Font(size = 15)
label_6.grid(row = 2, column = 0,sticky = "news", padx = 10)


label_7 = tk.Label(master = frame_3,relief = tk.RAISED, text = "No of days Absent:\n\t", bg = "lightblue", fg = 'red')
label_7['font'] = font.Font(size = 15)
label_7.grid(row = 2, column = 1 ,sticky = "news", padx = 10)

frame_4= tk.Frame(window,width  = 900, height = 100 , bg = "lightblue")
frame_4.pack()


frame_4.rowconfigure(0,weight = 1 , minsize = 45)
frame_4.columnconfigure(0,weight = 1 , minsize = 1000)

label_8 = tk.Label(master = frame_4,relief = tk.RIDGE, text = "ELIGIBILITY",bg = "white", fg = 'blue')
label_8['font'] = font.Font(size = 15)
label_8.grid(row = 0, column = 0 ,sticky = "news", padx = 50)


def view_attendance():
	print('Posting Attendance')
	os.system("""gnome-terminal -e " libreoffice /home/sammy-ros/attendance_facial/attendance.xlsx" """)


frame_5= tk.Frame(window,width  = 900, height = 100 , bg = "lightblue")
frame_5.pack()


frame_5.columnconfigure(1,weight = 1 , minsize = 500)
frame_5.columnconfigure(0,weight = 1 , minsize = 500)
frame_5.rowconfigure(0,weight = 1 , minsize = 100)



A = tk.Button(master = frame_5, text = 'View Spreadsheet',bg = "lightblue" , fg = "blue" ,cursor = "fleur",command =  view_attendance )
B = tk.Button(master = frame_5, text = 'Exit',bg = "lightblue" , fg = "blue",cursor = "fleur",command =  sys.exit )

A['font'] = font.Font(size = 18)
B['font'] = font.Font(size = 18)

A.grid(row=  0,column = 0, sticky = "nsew", padx = 35, pady = 15)
B.grid(row=  0,column = 1, sticky = "nsew", padx = 35, pady = 15)

window.mainloop()