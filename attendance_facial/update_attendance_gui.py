import sys

sys.path.append('~/.local/lib/python3.5/site-packages')                                     #for ros users
sys.path.remove('/opt/ros/kinetic/lib/python2.7/dist-packages')                             #for ros users

import time
import numpy as np
import matplotlib.pyplot as plt
import os
import cv2
from tqdm import tqdm
from numpy import argmax
import tkinter as tk
import tkinter.font as font



txtimg = "Path for test_image: "
txtvid = "Path for test_video: "

today_date = time.strftime('%d_%m_%y')

folderName = str(today_date)
rp = '/home/sammy-ros/attendance_facial'

folderPath = rp + "/attendace_data/" +folderName
if not os.path.exists(folderPath):
    os.makedirs(folderPath)



DATADIR = '/home/sammy-ros/attendance_facial/datasets'
CATEGORY=[]
for L in os.listdir(DATADIR): 
    CATEGORY.append(L)
#print(CATEGORY)


present_list = []
today_present  = []
today_absentees = []


face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
#face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_alt_tree.xml')


rec = cv2.face.LBPHFaceRecognizer_create()    
rec.read('trainingData.yml')                                       



boxlimit = 650
unk_samp = 0



from openpyxl import Workbook,load_workbook

file_name = "/home/sammy-ros/attendance_facial/attendance.xlsx"

if os.path.exists(file_name):
    wb = load_workbook(filename = file_name)
    sh1 = wb.active
    
else :
    print("create database before updating attendance")

Roll_No = []
for values in sh1.iter_cols(min_row = 2, min_col  = 1, max_col = 1, values_only = True):
    for x in values:
        if x != None : Roll_No.append(x)





from openpyxl import Workbook,load_workbook

file_name = "/home/sammy-ros/attendance_facial/attendance.xlsx"

if os.path.exists(file_name):
    wb = load_workbook(filename = file_name)
    sh1 = wb.active
    
else :
    print("create database before updating attendance")

Roll_No = []
for values in sh1.iter_cols(min_row = 2, min_col  = 1, max_col = 1, values_only = True):
    for x in values:
        if x != None : Roll_No.append(x)




def face_pred(img,frame,unk_samp):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)                             
    faces = face_cascade.detectMultiScale(gray)
    totalConf = 0.0
    faceRec = 0
    
    for (x, y, w, h) in faces:
        if True:

            folderPath_P = folderPath + '/' + 'predicted_faces'
            folderPath_U = folderPath + '/' + 'unknown_faces'
            if not os.path.exists(folderPath_P):
                os.makedirs(folderPath_P)
            if not os.path.exists(folderPath_U):
                os.makedirs(folderPath_U)



            cv2.rectangle(frame,(x,y),(w+x,y+h),(0,0,255),6)

            id, conf = rec.predict(gray[y:y+h, x:x+w])    # Comparing from the trained data
            
            if conf >= 50 and conf <=99:
                profile = CATEGORY[id-1]
                if profile not in present_list:
                    present_list.append(profile)
                    print(profile)
                cv2.putText(frame, profile,(x+w//2,y),cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,255),2,cv2.LINE_AA)
                
                 
                dfn = path = os.path.join(DATADIR,profile)
                fn = folderPath_P +'/' + profile +".jpg"
                cv2.imwrite( fn , frame[y:y+h, x:x+w])
                #cv2.imwrite( dfn+'/'+str(today_date) +'.jpg', frame[y:y+h, x:x+w])
            
            else :

                cv2.putText(frame, 'Unknown',(x+w//2,y),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,25),2,cv2.LINE_AA)          
                fn = folderPath_U +'/' +"Unknown"+ str(unk_samp) +".jpg"
                unk_samp+=1
                cv2.imwrite( fn , frame[y:y+h, x:x+w])
     
    return unk_samp,frame,present_list



video  = False



def video_load(cap,video,unk_samp):
    prev_val,prev_valu,present_list=[],[],[]
    present_roll,absent_roll = '',''
    while video:
        
        ret,frame = cap.read()
        if not ret:
            print("some problem occured")
            break
        frame=cv2.flip(frame,1)
        img = frame.copy()
        

        unk_samp,frame,present_list= face_pred(img,frame,unk_samp)
        cv2.putText(frame, 'No of present:'+str(len(present_list)),(10,470),cv2.FONT_HERSHEY_SIMPLEX,1,(200,255,25),2,cv2.LINE_AA)

        for x in present_list: 
            if x not in prev_val:
                present_roll = x  + ', ' +present_roll
                prev_val.append(x)
        

        label_5 = tk.Label(master = frame_3, text = present_roll,relief = tk.RIDGE, fg = 'red')
        label_5.grid(row = 1, column = 0,sticky = "nEws", padx = 25, pady = 0)

        

        cv2.imshow("OUTPUT",frame)
        k = cv2.waitKey(10)
        if k == ord('s'):
            break
    today_absentees = list(set(Roll_No) - set(present_list))
    for x in today_absentees: 
        if x not in prev_valu:
            absent_roll = x +', '+absent_roll 
            prev_valu.append(x)
    label_15 = tk.Label(master = frame_3, text = absent_roll,relief = tk.RIDGE, fg = 'red')
    label_15.grid(row = 1, column = 1,sticky = "nEws", padx = 25, pady = 0)
    label_02 = tk.Label(master = frame_2, text = str(len(present_list)), relief = tk.RAISED, bg = "lightblue", fg = 'red')
    label_02.grid(row = 1, column = 1,sticky = "news", padx = 5)

    label_03 = tk.Label(master = frame_2, text = str(len(today_absentees)), relief = tk.RAISED, bg = "lightblue", fg = 'red')
    label_03.grid(row = 1, column = 2,sticky = "news", padx = 5)


    if video :cap.release()
    cv2.destroyAllWindows()



def post_attendance():

    today_index = 0
    today_attendance  = []

    print('Posting Attendance')
    for values in sh1.iter_rows(min_row = 1, max_row =1, values_only = True):
        today_index  = len(values) + 1
        sh1.cell(row = 1, column = today_index).value = str(today_date)
        today_attendance = ['A'] * len(Roll_No)


    for present in present_list:
        if present in Roll_No:
            today_attendance[Roll_No.index(present)] = "P"
    print(today_attendance)

    for i in range(len(today_attendance)):
        sh1.cell(row = i+2, column = today_index ).value = today_attendance[i]

    wb.save(filename = file_name)


    
def cancel():
    
    wb.save(filename = file_name)
    import sys
    sys.exit()

    

def image_det():
    entry_1.delete(0,len(txtimg))
    img_path = entry_1.get()
    img_dir = img_path
    
    prev_val,prev_valu,present_list=[],[],[]
    present_roll,absent_roll = '',''
 
    if os.path.exists(img_dir):
        frame = cv2.imread(img_dir)
        
        img = frame.copy()
        
    
        unk_samp = 0

        unk_samp,frame,present_list  = face_pred(img,frame,unk_samp)

        
        for x in present_list: 
            if x not in prev_val:
                present_roll = x  + ', ' +present_roll
                prev_val.append(x)

        

        label_5 = tk.Label(master = frame_3, text = present_roll,relief = tk.RIDGE, fg = 'red')
        label_5.grid(row = 1, column = 0,sticky = "nEws", padx = 25, pady = 0)

        today_absentees = list(set(Roll_No) - set(present_list))
        for x in today_absentees: 
            if x not in prev_valu:
                absent_roll = x +', '+absent_roll 
                prev_valu.append(x)
    
            
        label_15 = tk.Label(master = frame_3, text = absent_roll,relief = tk.RIDGE, fg = 'red')
        label_15.grid(row = 1, column = 1,sticky = "nEws", padx = 25, pady = 0)
        label_02 = tk.Label(master = frame_2, text = str(len(present_list)), relief = tk.RAISED, bg = "lightblue", fg = 'red')
        label_02.grid(row = 1, column = 1,sticky = "news", padx = 5)

        label_03 = tk.Label(master = frame_2, text = str(len(today_absentees)), relief = tk.RAISED, bg = "lightblue", fg = 'red')
        label_03.grid(row = 1, column = 2,sticky = "news", padx = 5)
        
        cv2.destroyAllWindows()



def vid_det():
    entry_2.delete(0,len(txtvid))
    vid_dir = entry_2.get()
    
    if os.path.exists(vid_dir):
        video = True
        unk_samp = 0
        cap =cv2.VideoCapture(vid_dir)
        video_load(cap,video,unk_samp)


def live_stream():
    print("live straming")
    video = True
    unk_samp = 0
    cap =cv2.VideoCapture(-1)
    video_load(cap,video,unk_samp)



window = tk.Tk()
window.title("Update Attendance")
#____________________________________________________________________________________________
#frame 1 
#____________________________________________________________________________________________


frame_1 = tk.Frame(window,width  = 900, height = 100 )

frame_1.pack()

frame_1.rowconfigure(0 ,weight = 1 , minsize = 60)
for x in range(1,5):
    frame_1.rowconfigure(x,weight = 1 , minsize = 25)
frame_1.columnconfigure(0,weight = 1 , minsize = 600)
frame_1.columnconfigure(1,weight = 1 , minsize = 300)

label = tk.Label(master = frame_1, text = "Enter the file path for the test image or video.\n For live prediction Press the Live streaming Button", relief = tk.RAISED, fg = 'red')
label['font'] = font.Font(size = 13)
label.grid(row = 0, column = 0,sticky = "news", padx = 5)


entry_1 = tk.Entry(master = frame_1 )
entry_1.grid(row = 1, column = 0,sticky = "news", pady = 2)

entry_1.insert(0,txtimg)

entry_2 = tk.Entry(master = frame_1 )
entry_2.grid(row = 2, column = 0,sticky = "news", pady = 2)

entry_2.insert(0,txtvid)

A1 = tk.Button(master = frame_1, text = 'test image path',cursor = "fleur",bg = "lightblue" , fg = "blue" ,command =  image_det )
B1= tk.Button(master = frame_1, text = 'test video',bg = "lightblue" ,cursor = "fleur", fg = "blue",command =  vid_det )
C1= tk.Button(master = frame_1, text = 'live streaming',bg = "lightblue" ,cursor = "fleur", fg = "blue",command =  live_stream )
A1['font'] = font.Font(size = 12)
B1['font'] = font.Font(size = 12)
C1['font'] = font.Font(size = 13)
A1.grid(row=  1,column = 1, sticky = "nsew", padx = 5, pady = 2)
B1.grid(row=  2,column = 1, sticky = "nsew", padx = 5, pady = 2)
C1.grid(row=  3,column = 0, sticky = "nsew", padx = 5, pady = 8)


#____________________________________________________________________________________________
#frame 2 
#____________________________________________________________________________________________


frame_2 = tk.Frame(window,width  = 900, height = 100 , bg = "lightblue")
frame_2.pack()

for x in range(0,3):
    frame_2.columnconfigure(x,weight = 1 , minsize = 300)
for x in range(2):
    frame_2.rowconfigure(x,weight = 1 , minsize = 50)

label_1 = tk.Label(master = frame_2, text = "Total No of Students:", relief = tk.RAISED, bg = "lightblue", fg = 'red')
label_1['font'] = font.Font(size = 15)
label_1.grid(row = 0, column = 0,sticky = "news", padx = 5)


label_2 = tk.Label(master = frame_2, text = "No of Students Present:", relief = tk.RAISED, bg = "lightblue", fg = 'red')
label_2['font'] = font.Font(size = 15)
label_2.grid(row = 0, column = 1,sticky = "news", padx = 5)

label_3 = tk.Label(master = frame_2, text = "No of Students Absent:", relief = tk.RAISED, bg = "lightblue", fg = 'red')
label_3['font'] = font.Font(size = 15)
label_3.grid(row = 0, column = 2,sticky = "news", padx = 5)


label_01 = tk.Label(master = frame_2, text = str(len(Roll_No)), relief = tk.RAISED, bg = "lightblue", fg = 'red')
label_01['font'] = font.Font(size = 15)
label_01.grid(row = 1, column = 0,sticky = "news", padx = 5)


label_02 = tk.Label(master = frame_2, text = str(len(today_present)), relief = tk.RAISED, bg = "lightblue", fg = 'red')
label_02['font'] = font.Font(size = 15)
label_02.grid(row = 1, column = 1,sticky = "news", padx = 5)

label_03 = tk.Label(master = frame_2, text = str(len(today_absentees)), relief = tk.RAISED, bg = "lightblue", fg = 'red')
label_03['font'] = font.Font(size = 15)
label_03.grid(row = 1, column = 2,sticky = "news", padx = 5)


#____________________________________________________________________________________________
#frame 3
#____________________________________________________________________________________________


frame_3 = tk.Frame(window,width  = 900, height = 100, bg = "white" )
frame_3.pack()

for x in range(0,2):
    if x == 0:frame_3.rowconfigure(x,weight = 1 , minsize = 50)
    else:frame_3.rowconfigure(x,weight = 1 , minsize = 150)
    frame_3.columnconfigure(x,weight = 1 , minsize = 450)


label_4 = tk.Label(master = frame_3, text = "List of Students Present:", fg = 'red', bg = "white")
label_4['font'] = font.Font(size = 15)
label_4.grid(row = 0, column = 0,sticky = "ws", padx = 25, pady = 0)


label_5 = tk.Label(master = frame_3, text = "",relief = tk.RIDGE, fg = 'red')
label_5.grid(row = 1, column = 0,sticky = "nEws", padx = 25, pady = 0)

label_14 = tk.Label(master = frame_3, text = "List of students Absent :", fg = 'red', bg = "white")
label_14['font'] = font.Font(size = 15)
label_14.grid(row = 0, column = 1,sticky = "ws", padx = 25, pady = 0)


label_15 = tk.Label(master = frame_3, text = "",relief = tk.RIDGE, fg = 'red')

label_15.grid(row = 1, column = 1,sticky = "nEws", padx = 25, pady = 0)

#____________________________________________________________________________________________
#frame 4 
#____________________________________________________________________________________________

frame_4 = tk.Frame(window,width  = 900, height = 100 , bg = "white")
frame_4.pack()

frame_4.columnconfigure(1,weight = 1 , minsize = 450)
frame_4.columnconfigure(0,weight = 1 , minsize = 450)
frame_4.rowconfigure(0,weight = 1 , minsize = 100)


A = tk.Button(master = frame_4, text = 'Update Attendance',bg = "lightblue" , fg = "blue" ,cursor = "fleur",command =  post_attendance )
B = tk.Button(master = frame_4, text = 'Exit',bg = "lightblue" , fg = "blue",cursor = "fleur",command =  cancel )

A['font'] = font.Font(size = 18)
B['font'] = font.Font(size = 18)

A.grid(row=  0,column = 0, sticky = "nsew", padx = 35, pady = 15)
B.grid(row=  0,column = 1, sticky = "nsew", padx = 35, pady = 15)



window.mainloop()
