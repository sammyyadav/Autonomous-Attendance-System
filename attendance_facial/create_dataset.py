import sys
import cv2 
import matplotlib.pyplot as plt
import numpy as np
import time
import os

system_name = 'sammy-ros'

from openpyxl import Workbook,load_workbook

face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')                
#face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_alt_tree.xml')

file_name ='/home/'system_name'/attendance_facial/attendance.xlsx'

if os.path.exists(file_name):
    wb = load_workbook(filename = file_name)
    sh1 = wb.active
    
else:
    wb = Workbook()
    sh1 = wb.active
    sh1.title = "June"
    sh1.append( ("Roll No.", "Name"))

Roll_No = []
Name  = []

for values in sh1.iter_cols(min_row = 2, min_col  = 1, max_col = 1, values_only = True):
    for x in values:
        if x != None : Roll_No.append(x)
for values in sh1.iter_cols(min_row = 2, min_col  = 2, max_col = 2, values_only = True):
    for x in values:
        if x != None : Name.append(x)

New_Name = input("Enter the name of the student : ")
New_Roll_No = input("Enter the Roll Number of the student : ")


if New_Roll_No not in Roll_No:
    Roll_No.append(New_Roll_No)
if New_Name not in Name:
    Name.append(New_Name)

print (Roll_No)
print(Name)

for i in range(len(Roll_No)):
    sh1.cell(row = i+2, column = 2 ).value = Name[i]
    sh1.cell(row = i+2, column = 1).value = Roll_No[i]

print ('New data is added to the database')

prev_time = 0

folderName = New_Roll_No

folderPath = '/home/'system_name'/attendance_facial/'+ "datasets/"+folderName
if not os.path.exists(folderPath):
    os.makedirs(folderPath)
    

Num_of_datas = 0

cap =cv2.VideoCapture(-1)  

while True:
    
    ret,frame = cap.read()
    if ret == False:
        break
    frame=cv2.flip(frame,1)
    fac_img = frame.copy()
    fac_rects = face_cascade.detectMultiScale(fac_img)

    for (x,y,w,h) in fac_rects:
        
        cv2.rectangle(fac_img,(x,y),(w+x,y+h),(0,0,255),6)

        cur_time = time.time()
        dif_time = cur_time - prev_time

        if  dif_time > 0.75 and Num_of_datas < 20:
            fn = folderPath + '/' + New_Name + str(Num_of_datas) +".jpg"
            cv2.imwrite( fn , frame[y:y+h, x:x+w])
            prev_time = cur_time
            Num_of_datas += 1 

    if Num_of_datas == 20:
        break



    cv2.imshow("ss",fac_img)
    if cv2.waitKey(5) and 0xFF == 27:
        break

cap.release()
cv2.destroyAllWindows()

wb.save(filename = file_name)