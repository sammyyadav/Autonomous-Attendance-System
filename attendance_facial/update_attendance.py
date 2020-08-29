import sys

sys.path.append('~/.local/lib/python3.5/site-packages')                                     #for ros users
sys.path.remove('/opt/ros/kinetic/lib/python2.7/dist-packages')                             #for ros users

import time
import numpy as np
import matplotlib.pyplot as plt
import os
import cv2
from tqdm import tqdm
from numpy import argmax
from openpyxl import Workbook,load_workbook



today_date = time.strftime('%d_%m_%y')

folderName = str(today_date)
rp = '/home/sammy-ros/attendance_facial'

folderPath = rp + "/attendace_data/" +folderName
if not os.path.exists(folderPath):
    os.makedirs(folderPath)



DATADIR = '/home/sammy-ros/attendance_facial/datasets'
CATEGORY=[]
for L in os.listdir(DATADIR): 
    CATEGORY.append(L)
print(CATEGORY)


present_list = []


face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
#face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_alt_tree.xml')


rec = cv2.face.LBPHFaceRecognizer_create()    
rec.read('trainingData.yml')                                       



boxlimit = 650
unk_samp = 0


def face_pred(img,unk_samp):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)                             
    faces = face_cascade.detectMultiScale(gray)
    totalConf = 0.0
    faceRec = 0
    for (x, y, w, h) in faces:
    	if True:

    		folderPath_P = folderPath + '/' + 'predicted_faces'
    		folderPath_U = folderPath + '/' + 'unknown_faces'
    		if not os.path.exists(folderPath_P):
    			os.makedirs(folderPath_P)
    		if not os.path.exists(folderPath_U):
    			os.makedirs(folderPath_U)



    		cv2.rectangle(frame,(x,y),(w+x,y+h),(0,0,255),6)

    		id, conf = rec.predict(gray[y:y+h, x:x+w])    # Comparing from the trained data

    		if conf >= 50 and conf <=99:
    			profile = CATEGORY[id-1]
    			if profile not in present_list:
    				present_list.append(profile)
    				print(profile)
    			cv2.putText(frame, profile,(x+w//2,y),cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,255),2,cv2.LINE_AA)
    			
    			 
    			dfn = path = os.path.join(DATADIR,profile)
    			fn = folderPath_P +'/' + profile +".jpg"
    			cv2.imwrite( fn , frame[y:y+h, x:x+w])
    			#cv2.imwrite( dfn+'/'+str(today_date) +'.jpg', frame[y:y+h, x:x+w])
    		
    		else :

        		cv2.putText(frame, 'Unknown',(x+w//2,y),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,25),2,cv2.LINE_AA)        	
        		fn = folderPath_U +'/' +"Unknown"+ str(unk_samp) +".jpg"
        		unk_samp+=1
        		cv2.imwrite( fn , frame[y:y+h, x:x+w])
     
    return unk_samp



video  = False

print("Enter the path of the image or video if avialable.")      		

img_dir = input("\nEnter the path for the test image. \nFor entering test video path press Enter ")

if os.path.exists(img_dir):
	frame = cv2.imread(img_dir)
	frame = cv2.resize(frame, (640,480))
	img = frame.copy()
	
	print("\nList of students identified:")
	unk_samp  = face_pred(img,unk_samp)
	frame=cv2.flip(img,1)
	unk_samp  = face_pred(img,unk_samp)
	
	


	
elif img_dir == "":
	vid_dir = input("\nEnter the path for the test video. \nFor entering live video prediction press Enter ")
	print("\npress 's' to show attendance :")
	print("\nList of students identified:")
	if os.path.exists(vid_dir):
		video = True
		cap =cv2.VideoCapture(vid_dir)

	elif vid_dir =='':
		video = True
		cap =cv2.VideoCapture(-1)
	
    


while video:
    ret,frame = cap.read()
    if not ret:break
    frame=cv2.flip(frame,1)
    img = frame.copy()
    

    unk_samp = face_pred(img,unk_samp)
    cv2.putText(frame, 'No of present:'+str(len(present_list)),(10,470),cv2.FONT_HERSHEY_SIMPLEX,1,(200,255,25),2,cv2.LINE_AA)
    cv2.imshow("OUTPUT",frame)
    k = cv2.waitKey(10)
    if k == ord('s'):
        break

if img_dir == '':cap.release()
cv2.destroyAllWindows()

if unk_samp != 0:
	print("there are ", unk_samp, " unknown faces.\n Do you want to add these faces to datasets?")
	z = input("Enter Y/N :")
	
	if z.lower() == "y":
		samp_no = 0
		unk_samp_path = folderPath + '/unknown_faces'
		for img in tqdm(os.listdir(unk_samp_path)):
		
			img_array = cv2.imread(os.path.join(unk_samp_path,img) ) 
			while True:
				cv2.imshow("unknown",img_array)
				k = cv2.waitKey(10)
				if k == ord('s'): 
					Name = input("enter the name")
					Roll = input("enter the Roll_No")
					fn = DATADIR + '/'+ Roll
					if not os.path.exists(fn):
						os.makedirs(fn)
					cv2.imwrite( fn +'/'+ Name + str(samp_no)+".jpg" , img_array)
				elif k == ord('n'):
					break
		cv2.destroyAllWindows()
		


file_name = "/home/sammy-ros/attendance_facial/attendance.xlsx"

if os.path.exists(file_name):
    wb = load_workbook(filename = file_name)
    sh1 = wb.active
    
else :
	print("create database before updating attendance")

Roll_No = []
for values in sh1.iter_cols(min_row = 2, min_col  = 1, max_col = 1, values_only = True):
    for x in values:
        if x != None : Roll_No.append(x)



today_index = 0
today_attendance  = []
today_present =present_list
today_absentees = list(set(Roll_No) - set(today_present))


print('total stregnth  : ', len(Roll_No), "\n no of present : ", len(today_present),
						 "\t No of absentees : ",  len(today_absentees) )


print('''\ntoday's absentees are :''' )
for values in today_absentees:
	print("\t",values) 

x = input("verify the attendance. \n \n to update the attendance enter Y/y.")

if x == 'n' or x=='N':

	wb.save(filename = file_name)

def post_attendance(today_present):
	for values in sh1.iter_rows(min_row = 1, max_row =1, values_only = True):
		today_index  = len(values) + 1
		sh1.cell(row = 1, column = today_index).value = str(today_date)
		today_attendance = ['A'] * len(Roll_No)

	for present in today_present:
		if present in Roll_No:
			today_attendance[Roll_No.index(present)] = "P"

	for i in range(len(today_attendance)):
	    sh1.cell(row = i+2, column = today_index ).value = today_attendance[i]

	wb.save(filename = file_name)


if x == "y" or x=="Y":
	post_attendance(today_present)



	for values in sh1.iter_rows(min_row = 1, max_row =1, values_only = True):
		today_index  = len(values) + 1
		sh1.cell(row = 1, column = today_index).value = str(today_date)
		today_attendance = ['A'] * len(Roll_No)




