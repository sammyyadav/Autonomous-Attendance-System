import sys

sys.path.append('~/.local/lib/python3.5/site-packages')                                     #for ros users
sys.path.remove('/opt/ros/kinetic/lib/python2.7/dist-packages')                             #for ros users

import cv2 
import matplotlib.pyplot as plt
import numpy as np
import time
import os
import tkinter as tk
import tkinter.font as font
from openpyxl import Workbook,load_workbook

face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')                


def create_dataset():

	New_Name = entry_1.get()
	New_Roll_No  = entry_2.get()
	

	prev_time = 0

	folderName = New_Roll_No

	folderPath = '/home/sammy-ros/attendance_facial/'+ "datasets/"+folderName
	if not os.path.exists(folderPath):
	    os.makedirs(folderPath)
 
	Num_of_datas = 0

	cap =cv2.VideoCapture(-1)  

	while True:
	    
	    ret,frame = cap.read()
	    if ret == False:
	        break
	    frame=cv2.flip(frame,1)
	    fac_img = frame.copy()
	    fac_rects = face_cascade.detectMultiScale(fac_img)

	    for (x,y,w,h) in fac_rects:
	        
	        cv2.rectangle(fac_img,(x,y),(w+x,y+h),(0,0,255),6)

	        cur_time = time.time()
	        dif_time = cur_time - prev_time

	        if  dif_time > 0.75 and Num_of_datas < 20:
	            fn = folderPath + '/' + New_Name + str(Num_of_datas) +".jpg"
	            cv2.imwrite( fn , frame[y:y+h, x:x+w])
	            prev_time = cur_time
	            Num_of_datas += 1 

	    if Num_of_datas == 20:
	        break



	    cv2.imshow("ss",fac_img)
	    if cv2.waitKey(5) and 0xFF == 27:
	        break

	cap.release()
	cv2.destroyAllWindows()


def update_spreadsheet():
	file_name ='/home/sammy-ros/attendance_facial//attendance.xlsx'

	if os.path.exists(file_name):
	    wb = load_workbook(filename = file_name)
	    sh1 = wb.active
	    
	else:
	    wb = Workbook()
	    sh1 = wb.active
	    sh1.title = "June"
	    sh1.append( ("Roll No.", "Name"))

	Roll_No = []
	Name  = []

	for values in sh1.iter_cols(min_row = 2, min_col  = 1, max_col = 1, values_only = True):
	    for x in values:
	        if x != None : Roll_No.append(x)
	for values in sh1.iter_cols(min_row = 2, min_col  = 2, max_col = 2, values_only = True):
	    for x in values:
	        if x != None : Name.append(x)

	New_Name = entry_1.get()
	New_Roll_No  = entry_2.get()
	print(New_Name,New_Roll_No)

	if New_Roll_No not in Roll_No:
	    Roll_No.append(New_Roll_No)
	if New_Name not in Name:
	    Name.append(New_Name)

	
	for i in range(len(Roll_No)):
	    
	    sh1.cell(row = i+2, column = 2 ).value = Name[i]
	    sh1.cell(row = i+2, column = 1).value = Roll_No[i]

	wb.save(filename = file_name)


	


def train_dataset():
	print('Training Dataset')
	os.system("""gnome-terminal -e "python3 /home/sammy-ros/attendance_facial/training.py" """)


def cancel():
    print('exiting')
    import sys
    sys.exit()


window = tk.Tk()
window.title('Create Dataset')

frame_1 = tk.Frame(window,width  = 900, height = 100 , bg = "lightgreen")
frame_1.pack()


frame_1.rowconfigure(0,weight = 1 , minsize = 25)
frame_1.columnconfigure(0,weight = 1 , minsize = 900)


label = tk.Label(master = frame_1, text = "Create dataset", bg = "lightgreen", fg = 'red')
label['font'] = font.Font(family = "Times",size = 30)
label.grid(row = 0, column = 0,sticky = "news", padx = 5)




frame_2 = tk.Frame(window,width  = 1900, height = 100 , bg = "lightblue")
frame_2.pack()

frame_2.columnconfigure(0 ,weight = 1 , minsize = 900)
for x in range(4):
	frame_2.rowconfigure(x,weight = 1 , minsize = 50)

label_1 = tk.Label(master = frame_2, text = "Enter the Name:", bg = "lightblue", fg = 'red')
label_1['font'] = font.Font(size = 15)
label_1.grid(row = 0, column = 0,sticky = "ws", padx = 50)


label_2 = tk.Label(master = frame_2, text = "Enter the Roll Number:", bg = "lightblue", fg = 'red')
label_2['font'] = font.Font(size = 15)
label_2.grid(row = 2, column = 0 ,sticky = "ws", padx = 50)



entry_1 = tk.Entry(master = frame_2 )
entry_1['font'] = font.Font(size = 15)
entry_1.grid(row = 1, column = 0,sticky = "news",padx = 50)


entry_2 = tk.Entry(master = frame_2 )
entry_2['font'] = font.Font(size = 15)
entry_2.grid(row = 3, column = 0,sticky = "news",padx = 50)




frame_4 = tk.Frame(window,width  = 900, height = 100 , bg = "lightblue")
frame_4.pack()
for x in range(2):
	frame_4.columnconfigure(x,weight = 1 , minsize = 450)
	frame_4.rowconfigure(x,weight = 1 , minsize = 100)

A = tk.Button(master = frame_4, text = 'Create Dataset',cursor = "fleur",bg = "lightblue" , fg = "blue" ,command =  create_dataset )
B = tk.Button(master = frame_4, text = 'Train Dataset',cursor = "fleur",bg = "lightblue" , fg = "blue",command =  train_dataset )

D = tk.Button(master = frame_4, text = 'Update Spreadsheet',bg = "lightblue" , fg = "blue" ,command =  update_spreadsheet )
D['font'] = font.Font(size = 18)
D.grid(row=  0,column = 1, sticky = "nsew", padx = 35, pady = 15)

A['font'] = font.Font(size = 18)
B['font'] = font.Font(size = 18)

A.grid(row=  0,column = 0, sticky = "nsew", padx = 35, pady = 15)
B.grid(row=  1,column = 0, sticky = "nsew", padx = 35, pady = 15)


C= tk.Button(master = frame_4, text = 'Exit',cursor = "fleur",bg = "lightblue" , fg = "blue",command =  cancel )
C['font'] = font.Font(size = 18)
C.grid(row=  1,column = 1, sticky = "nsew", padx = 35, pady = 18)





window.mainloop()